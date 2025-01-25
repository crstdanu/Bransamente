import os
import json
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def content_update():
    # Update these paths
    # Path to the Excel file
    excel_file_path = "G:/Shared drives/Root/04. Bransamente RGT/Pentru Misa.xlsx"
    # Path to store folder content information
    content_record_path = "directory_content.json"
    # Define the base directory path
    folder_location = "G:/Shared drives/Root/04. Bransamente RGT"

    # Load the workbook and select the active worksheet
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # Load the previously stored directory content data
    if os.path.exists(content_record_path):
        with open(content_record_path, "r") as file:
            previous_content = json.load(file)
    else:
        previous_content = {}

    # Define a green fill color for modified directories
    green_fill = PatternFill(start_color="00FF00",
                             end_color="00FF00", fill_type="solid")

    # Iterate through folder names in the Excel file
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        folder_name = row[0].value
        if not folder_name:
            continue

        folder_path = os.path.join(folder_location, folder_name)

        if not os.path.exists(folder_path):
            print(f"Folder not found: {folder_path}")
            continue

        # Get current content of the folder
        current_content = {}
        for root, dirs, files in os.walk(folder_path):
            for name in dirs + files:
                item_path = os.path.join(root, name)
                current_content[item_path] = datetime.fromtimestamp(
                    os.path.getmtime(item_path)).isoformat()

        # Compare with previous content
        folder_modified = False
        if folder_name in previous_content:
            if current_content != previous_content[folder_name]:
                folder_modified = True
        else:
            folder_modified = True

        # Update the Excel sheet if changes are detected
        if folder_modified:
            row[0].fill = green_fill  # Highlight the folder name in green

        # Update the recorded content
        previous_content[folder_name] = current_content

    # Save the updated content to the JSON file
    with open(content_record_path, "w") as file:
        json.dump(previous_content, file, indent=4)

    # Save the updated workbook
    workbook.save(excel_file_path)

    print("Directory content check complete.")

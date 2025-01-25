import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill


def folder_list():
    # Update these paths
    # Replace with the path to your Excel file
    excel_file_path = "G:/Shared drives/Root/04. Bransamente RGT/Pentru Misa.xlsx"
    # Replace with the location of folders on your drive
    folder_location = "G:/Shared drives/Root/04. Bransamente RGT"

    # Load the workbook and select the active worksheet
    workbook = load_workbook(excel_file_path)
    sheet = workbook.active

    # Read the existing folder names from the second column of the Excel sheet
    existing_folders = set()
    for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=2, max_col=2):
        folder_name = row[0].value
        if folder_name:
            existing_folders.add(folder_name)

    # Get the list of folder names from the specified directory
    current_folders = {item for item in os.listdir(
        folder_location) if os.path.isdir(os.path.join(folder_location, item))}

    # Find new folders that are not already in the Excel sheet
    new_folders = current_folders - existing_folders

    # Define a red fill color for new cells
    red_fill = PatternFill(start_color="FF0000",
                           end_color="FF0000", fill_type="solid")

    # Write the new folders to the next available rows in the Excel sheet
    next_row = 2  # Start after the header row
    for folder_name in new_folders:
        # Find the next blank cell in column 2
        while sheet.cell(row=next_row, column=2).value:
            next_row += 1
        # Add a sequential number in the first column
        sheet.cell(row=next_row, column=1, value=next_row - 1)
        # Add the folder name in the second column
        cell = sheet.cell(row=next_row, column=2, value=folder_name)
        cell.fill = red_fill  # Set the cell color to red

    # Save the updated workbook
    workbook.save(excel_file_path)

    print(f"Am adaugat {len(new_folders)} directoare in Excel.")
